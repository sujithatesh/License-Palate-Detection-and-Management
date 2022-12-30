import tkinter as tk
from tkinter import *
import openpyxl
import torch
from matplotlib import pyplot as plt
import numpy as np
import cv2
import os
import enchant
from easyocr import Reader
from datetime import datetime
from playsound import playsound
def MainWindow():
    global mainWindow
    mainWindow = Tk()
    mainWindow.geometry('400x400')

    homef=Frame(mainWindow)
    home = Label(homef,text="HOME")
    home.config(font=('calibri 40'))
    home.pack()


    whitespace=Label(homef,text="  ")
    whitespace.config(font=('calibri 40'))
    whitespace.pack()

    homef.pack()
    btn1=Frame(mainWindow)
    def status():
        mainWindow.destroy()
        call_status()

    def detect():
        mainWindow.destroy()
        call_detect()
    def registerb():
        mainWindow.destroy()
        registers()
        
    button1=Button(homef,text="Check the status",command=status)
    button1.config(font=('calibri 20'))
    button1.pack()
    #btn1.pack()

    whitespace1 = Label(homef, text="  ")
    whitespace1.config(font=('calibri 10'))
    whitespace1.pack()

    button2=Button(mainWindow,text="Detect",command=detect)
    button2.config(font=('calibri 20'))
    button2.pack()

    
    whitespace2 = Label(mainWindow, text="  ")
    whitespace2.config(font=('calibri 10'))
    whitespace2.pack()

    register=Button(mainWindow,text="Register a new vehicle",command=registerb)
    register.config(font=('calibri 20'))
    register.pack()
    
    mainWindow.mainloop()
def registers():
    r=Tk()
    r.geometry('800x400')

    title=Label(r,text="Registration")
    title.config(font=('calibri 30'))
    title.pack()

    whitespace=Label(r,text="   ")
    whitespace.config(font=('calibri 30'))
    whitespace.pack()
    
    entryf = Frame(r)
    l=Label(entryf,text="Enter the vehicle registration plate number :")
    l.config(font=('calibri 20'))
    l.pack(side=LEFT)

    e=Entry(entryf)
    e.config(font=('calibri 20'))
    e.pack(side=RIGHT)

    entryf.pack()

    whitespace1=Label(r,text="   ")
    whitespace1.config(font=('calibri 20'))
    whitespace1.pack()

    def backb():
        r.destroy()
        MainWindow()
    def regd():
        wb=openpyxl.load_workbook('data.xlsx')
        sh1=wb['Sheet1']
        mrows=sh1.max_row
        num=e.get()
        num=num.upper()
        if num not in numberplates:
            sh1.cell(row=mrows+1,column=1).value=num
            numberplates.append(num)
            ans.config(text="Successfully resgistered")
            ans.pack()
        else:
            ans.config(text="Already registered")
            ans.pack()
        wb.save("data.xlsx")
         

    ansf=Frame(r)
    ans=Label(ansf)
    ans.config(font=('calibri 20'))
    reg=Button(ansf,text="Register",command=regd)
    reg.config(font=('calibri 20'))
    reg.pack()

    ansf.pack()
    whitespace3=Label(r,text="   ")
    whitespace3.config(font=('calibri 40'))
    whitespace3.pack()

    backframe=Frame(r)
    whitespace2=Label(backframe,text="  ")
    whitespace2.config(font=('calibri 20'))
    whitespace2.pack(side=LEFT)
    back = Button(backframe,text="Back",command=backb)
    back.config(font=('calibri 20'))
    back.pack(side=LEFT)

    backframe.pack(side=LEFT)
    
    
    
def call_status():
    update()
    global status
    status = Tk()
    status.geometry('500x400')

    title=Label(status,text="Check the status of a Car")
    title.config(font=('calibri 30'))
    title.pack()

    whitespace=Label(status,text="")
    whitespace.config(font=('calibri 30'))
    whitespace.pack()

    textf=Frame(status)

    nplate=Label(textf,text="Enter the number plate:")
    nplate.config(font=('calibri 15'))
    nplate.pack(side=LEFT)

    platentry=Entry(textf)
    platentry.config(font=('calibri 15'))
    platentry.pack(side=RIGHT)

    textf.pack()

    whitespace1 = Label(status, text="")
    whitespace1.config(font=('calibri 30'))
    whitespace1.pack()

    ansframe=Frame(status)
    ans=Label(ansframe)
    ans.config(font=('calibri 30'))

    def checknp():
        number = platentry.get()
        number=number.upper()
        print(number)
        print(inside)
        if number in numberplates:
            if number in inside:
                ans.config(text="Inside")
                ans.pack()
            else:
                ans.config(text="Outside")
                ans.pack()
        else:
            ans.config(text="Vehicle not registered")
            ans.pack()

    def backb():
        status.destroy()
        MainWindow()

    
    check=Button(ansframe,text="Check",command=checknp)
    check.config(font=('calibri 20'))
    check.pack()

    ansframe.pack()
    whitespace1 = Label(status, text="")
    whitespace1.config(font=('calibri 20'))
    whitespace1.pack()

    backframe=Frame(status)
    whitespace2=Label(backframe,text="  ")
    whitespace2.config(font=('calibri 20'))
    whitespace2.pack(side=LEFT)
    back = Button(backframe,text="Back",command=backb)
    back.config(font=('calibri 20'))
    back.pack(side=LEFT)

    backframe.pack(side=LEFT)
    


    status.mainloop()

def call_detect():
    global detect
    detect=Tk()

    detect.geometry('400x400')

    title=Label(detect,text="DETECTION")
    title.config(font=('calibri 30'))
    title.pack()
    
    whitespace1 = Label(detect, text="")
    whitespace1.config(font=('calibri 30'))
    whitespace1.pack()
    def entry():
        entry_detection()
    def exitg():
        exit_detection()

    def backb():
        detect.destroy()
        MainWindow()
        
    en = Button(detect,text="ENTRY GATE",command=entry)
    en.config(font=('calibri 20'))
    en.pack()

    whitespace = Label(detect, text="")
    whitespace.config(font=('calibri 20'))
    whitespace.pack()
    
    ex = Button(detect,text="EXIT GATE",command=exitg)
    ex.config(font=('calibri 20'))
    ex.pack()

    backframe=Frame(detect)
    whitespace2=Label(backframe,text="  ")
    whitespace2.config(font=('calibri 20'))
    whitespace2.pack(side=LEFT)
    back = Button(backframe,text="Back",command=backb)
    back.config(font=('calibri 20'))
    back.pack(side=LEFT)
    backframe.pack(side=LEFT)

    
def accurate(actual):
    for i in numberplates:
        if(enchant.utils.levenshtein(actual, i) <=2):
            return i
        
def exit_detection():
    def most_common(lst):
        return max(set(lst), key=lst.count, default=None)

    cap = cv2.VideoCapture(0)
    reader = Reader(['en'])
    mylist = []
    flag = False
    last=""
    while cap.isOpened():
        ret, frame = cap.read()

        txt = ""

        # now = datetime.now()
        # Make detections
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
        if ret == False:
            break
        results = model(frame)
        cv2.imshow('YOLO', np.squeeze(results.render()))
        if not results:
            continue
        imgt = results.xyxy[0]
        i = list(imgt)
        l = []
        for j in i:
            temp = j.cpu().numpy()
            l.append(temp)
        if not l or len(l[0]) < 4:
            continue
        m = l[0][4]
        pos = 0
        for j in range(1, len(l)):
            if l[j][4] > m:
                m = l[j][4]
                pos = i
        if (m < 0.9):
            continue
        xmin = int(l[pos][0])
        ymin = int(l[pos][1])
        xmax = int(l[pos][2])
        ymax = int(l[pos][3])
        # img1=mpimg.imread(frame)
        modimg = frame[ymin:ymax, xmin:xmax, :]
        ans = reader.readtext(modimg)

        # print(ans)
        if (not ans or len(ans) < 2 or len(ans[1]) < 2):
            continue
        else:
            for i in range(len(ans)):
                if ans[i][1] == 'IND' or 'ND' in ans[i][1]:
                    continue
                txt = txt + ans[i][1]
            if (flag == False):
                mylist.append(txt)
            # print("\n\n")
            # print(mylist)
        # print(results)
        if (len(mylist) > 30 and flag==False):
            actual = most_common(mylist)
            actual = accurate(actual)
            flag = True
            print(actual)
            wb=openpyxl.load_workbook('data.xlsx')
            sh1=wb['Sheet1']
            mrows=sh1.max_row
            pos=1
            if last == actual :
                continue
            for i in range(2,mrows+1):
                if sh1.cell(row=i,column=1).value==actual:
                    pos=i
                    break
            if pos==1:
                continue
            
            if sh1.cell(row=pos,column=4).value=="OUTSIDE":
               global alarm
               alarm = Tk()
               #alarm.config(bg='red')
               l = Label(alarm,text="DUPLICATE VEHICLE")
               #l.config(bg='white')
               l.config(font=('calibri 40'))
               l.pack()
               playsound('Music1.mp3')

               break
            else:
               sh1.cell(row=pos,column=3).value=datetime.now().strftime("%H:%M:%S")
               sh1.cell(row=pos,column=4).value="OUTSIDE" 
               #outside.append(actual)
               #inside.remove(actual)
               last=actual
               update()
            wb.save("data.xlsx")                
            #update()
            
                          
                    
        if (flag == True):
            if (enchant.utils.levenshtein(actual, txt) > 5):
                mylist = []
                flag = False
        
    # print(i)
    cap.release()
    cv2.destroyAllWindows()
    
def entry_detection():
    def most_common(lst):
        return max(set(lst), key=lst.count, default=None)

    cap = cv2.VideoCapture(0)
    reader = Reader(['en'])
    mylist = []
    flag = False
    last=""
    while cap.isOpened():
        ret, frame = cap.read()

        txt = ""

        # now = datetime.now()
        # Make detections
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
        if ret == False:
            break
        results = model(frame)
        cv2.imshow('YOLO', np.squeeze(results.render()))
        if not results:
            continue
        imgt = results.xyxy[0]
        i = list(imgt)
        l = []
        for j in i:
            temp = j.cpu().numpy()
            l.append(temp)
        if not l or len(l[0]) < 4:
            continue
        m = l[0][4]
        pos = 0
        for j in range(1, len(l)):
            if l[j][4] > m:
                m = l[j][4]
                pos = i
        if (m < 0.9):
            continue
        xmin = int(l[pos][0])
        ymin = int(l[pos][1])
        xmax = int(l[pos][2])
        ymax = int(l[pos][3])
        # img1=mpimg.imread(frame)
        modimg = frame[ymin:ymax, xmin:xmax, :]
        ans = reader.readtext(modimg)

        # print(ans)
        if (not ans or len(ans) < 2 or len(ans[1]) < 2):
            continue
        else:
            for i in range(len(ans)):
                if ans[i][1] == 'IND' or 'ND' in ans[i][1]:
                    continue
                txt = txt + ans[i][1]
            if (flag == False):
                mylist.append(txt)
            # print("\n\n")
            # print(mylist)
        # print(results)
        if (len(mylist) > 30 and flag == False):
            actual = most_common(mylist)
            actual = accurate(actual)
            flag = True
            print(actual)
            wb=openpyxl.load_workbook('data.xlsx')
            sh1=wb['Sheet1']
            mrows=sh1.max_row
            pos=1
            if last==actual:
                mylist=[]
                flag=False
                continue
            for i in range(2,mrows+1):
                if sh1.cell(row=i,column=1).value==actual:
                    pos=i
                    break
            if pos==1 :
                global notfound
                notfound = Tk()
                notfound.config(bg='red')
                l=Label(notfound,text="VEHICLE NOT FOUND")
                #l.config(bg='black')
                l.config(font=('calibri 30'))
                l.pack()
                playsound('Music1.mp3')
                break
            else:
                if sh1.cell(row=pos,column=4).value=="INSIDE":
                    global alarm
                    alarm = Tk()
                    alarm.config(bg='RED')
                    l = Label(alarm,text="DUPLICATE VEHICLE TRYING TO ENTER")
                    #l.config(bg='BLACK')
                    l.config(font=('calibri 40'))
                    l.pack()
                    playsound('Music1.mp3')
                    break
                else:
                    sh1.cell(row=pos,column=2).value=datetime.now().strftime("%H:%M:%S")
                    sh1.cell(row=pos,column=4).value="INSIDE"
                    #outside.remove(actual)
                    #inside.append(actual)
                    last=actual
                    update()
            
            wb.save("data.xlsx")
            #update()
                    
        if (flag == True):
            if (enchant.utils.levenshtein(actual, txt) > 5):
                mylist = []
                flag = False

    # print(i)
    cap.release()
    cv2.destroyAllWindows()
def update():
    inside.clear()
    outside.clear()
    wb=openpyxl.load_workbook('data.xlsx')
    sh1=wb['Sheet1']
    mrows=sh1.max_row
    for i in range(2,mrows+1):
        numberplates.append(sh1.cell(row=i,column=1).value)
        if sh1.cell(row=i,column=4).value=="INSIDE":
            inside.append(sh1.cell(row=i,column=1).value)
        else:
            outside.append(sh1.cell(row=i,column=1).value)

                            
    
    
    
if __name__=='__main__':
    model = torch.hub.load('ultralytics/yolov5', 'custom', path='C:/Users/sujit/Sujith/Mini Project 2/Project final/yolov5/runs/train/exp6/weights/last.pt', force_reload=True)
    inside=[]
    outside=[]
    numberplates=[]
    update()
    MainWindow()
